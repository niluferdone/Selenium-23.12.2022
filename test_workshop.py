import pytest
import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from constants import *
import openpyxl

class Test_Workshop():

  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(BASE_DOMAIN_URL)
    self.driver.maximize_window()

  def element_visible(self,id,value):
      WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((id, value)))
  
  def teardown_method(self):
    self.driver.quit()
  
  def readProductListFromExcel():
    # exceli aç
    # hücreleri oku 
    # okuduğun tüm verileri bir listeye at
    # listeyi return et
    excelFile = openpyxl.load_workbook("data/productList.xlsx")
    selectedSheet = excelFile["Sheet1"] # excel dosyasının Sheet1 isimli sayfasını çağırdık 

    rows = selectedSheet.max_row # kaç satır varsa
    productList = [] # verileri tutacak boş bir liste oluştur


    for i in range(2, rows+1): # 2 => 1.satırı başlık olarak görsün diyerek, rows+1 => 0 dan başladığı için son satırı da yazsın diye 
        productName = selectedSheet.cell(i,2).value  # productName 2.sütundan başlıyor
        tupleExample = (productName)
        productList.append(tupleExample)
    return productList 



  # 1- Normal giriş yapılması
  def test_login(self):
    
    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_INFO)))
    username= self.driver.find_element(By.ID, USERNAME_INFO)
    username.send_keys(USERNAME)

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_INFO)))
    # Password: secret_sauce
    password= self.driver.find_element(By.ID, PASSWORD_INFO)
    password.send_keys(PASSWORD)

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_INFO)))
    loginBtn= self.driver.find_element(By.ID, LGN_BUTTON)
    loginBtn.click()

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, MENU)))
    menu= self.driver.find_element(By.ID, MENU)
    menuText= menu.text

    assert menuText == MENU_TEXT

  # 1.1- ürünlerin sayısının doğrulanması
    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, PRODUCT_LIST)))
    productList = self.driver.find_elements(By.XPATH, PRODUCT_LIST)
    productListLen = len(productList)

    assert productListLen == 6



  # 2- "locked_out_user" ile giriş yapıldığında verilen uyarı mesajının doğrulanması.
  def test_login_fail(self): 

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_INFO)))
    username= self.driver.find_element(By.ID, USERNAME_INFO)
    username.send_keys(USERNAME2)

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_INFO)))
    password= self.driver.find_element(By.ID, PASSWORD_INFO)
    password.send_keys(PASSWORD)

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_INFO)))
    loginBtn= self.driver.find_element(By.ID, LGN_BUTTON)
    loginBtn.click()

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, ERROR)))
    error= self.driver.find_elements(By.XPATH, ERROR)
    errorLen = len(error)

    assert errorLen > 0 

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, ERROR)))
    error= self.driver.find_element(By.XPATH, ERROR)
    errorText = error.text

    assert errorText == ERROR_TEXT2

    

  # 3 - Ürünlerin isimlerinin excel dosyalarındaki isimlerle uyuşması 
  def test_checkingNames(self):
        excelFile = openpyxl.load_workbook(PRODUCT_NAMES_LİST)
        selectedSheet = excelFile[SHEET]
        rows = selectedSheet.max_row
        data = []
        for i in range(2,rows+1):
            productname = selectedSheet.cell(i,1).value
            data.append(productname)

        self.test_login()
        products = self.driver.find_elements(By.XPATH, PRODUCT_NAME)
        pnames = []
        
        for i in range(len(products)):
            product = products[i]
            productNamesText = product.text
            pnames.append(productNamesText)
        assert pnames[0] == data[0]

  # 4- Ürünlerin z'den a ya sıralanma fonksiyonun test edilmesi
  def test_product_desc_char(self):
    self.test_login()

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, ALIGNMENT)))
    product_alignment = self.driver.find_element(By.CLASS_NAME, ALIGNMENT)
    product_alignment.click()

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, SET_DESC_CHAR)))
    product_desc_char = self.driver.find_element(By.XPATH, SET_DESC_CHAR)
    product_desc_char.click()

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, PRODUCT_LIST)))
    product = self.driver.find_elements(By.XPATH, PRODUCT_LIST)
    productListName = self.driver.find_elements(By.XPATH, PRODUCT_NAME) # isimleri listeye at
    productNameLen = len(product) # product sayısı kadar 
    
    productList = []
    for i in range (0, productNameLen): # product sayısı kadar dön
        product_name = productListName[i] # ürünlerin indis değeri kadar atama yap 
        productText = product_name.text # prod un text değerini ata
        productList.append(productText) # ürün listesine prod un text değerini ata
        productList = sorted(productList, reverse = True)       
        #print(productList)
    return productList

  

  # 5- Ürünlerin düşük fiyattan yüksek fiyata sıralanma fonksiyonunun testi
  def test_product_asc_price(self):
    self.test_login()

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, ALIGNMENT)))
    product_alignment = self.driver.find_element(By.CLASS_NAME, ALIGNMENT)
    product_alignment.click()

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, SET_ASC_PRICE)))
    product_asc_price = self.driver.find_element(By.XPATH, SET_ASC_PRICE)
    product_asc_price.click()

    WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, PRODUCT_LIST)))
    price = self.driver.find_elements(By.XPATH, PRODUCT_LIST)
    productPrice = self.driver.find_elements(By.XPATH, PRODUCT_PRICE) 
    productPriceLen = len(price)
    
    productList = []
    for i in range (0, productPriceLen):
        product_price = productPrice[i] 
        product_id = float(product_price.text.replace("$",""))
        productList.append(product_id)
        productList=sorted(productList)
        #print(productList)
    return productList


  
  # 6- Bir excel dosyasında ismi geçen ürünlerin sepete eklenmesi fonksiyonu testi
  def test_product_add_basket(self):
    self.test_login()
    excelFile = openpyxl.load_workbook(PRODUCTS_ON_BASKET)
    selectedSheet = excelFile[SHEET]
    rows = selectedSheet.max_row

    data = []
    for i in range(2,rows+1):
        productName = selectedSheet.cell(i,1).value
        data.append(productName)

    if productName == FIRST_PRODUCT: 
        addBasket = self.driver.find_element(By.ID, FIRST_PRODUCT_ID)
        addBasket.click()
        basket = self.driver.find_elements(By.CLASS_NAME,BASKET_LINK)
        pOnBasket = len(basket)
        assert pOnBasket > 0

    elif productName == SECOND_PRODUCT: 
        addBasket = self.driver.find_element(By.ID, SECOND_PRODUCT_ID)
        addBasket.click()
        basket = self.driver.find_elements(By.CLASS_NAME,BASKET_LINK)
        pOnBasket = len(basket)
        assert pOnBasket > 0




  
  # 7- Sepete eklenen ürünlerin sepet sayfasında doğru bir şekilde görünmesi testi
  def test_product_to_basket(self):
    self.test_login()
    
    self.element_visible(By.ID, ADD_TO_CARD_PRODUCT_NAME_6)
    addToCard = self.driver.find_element(By.ID, ADD_TO_CARD_PRODUCT_NAME_6)
    addToCard.click()

    self.element_visible(By.XPATH, BASKET_CHECK)
    basketItem = self.driver.find_elements(By.XPATH, BASKET_CHECK)
    basketLen = len(basketItem)
        
    assert basketLen > 0 



  # 8- Sepetten kaldırılan ürünün sepet ekranından kaldırılma test
  def test_remove_product_from_basket(self):
    self.test_login()

    self.driver.find_element(By.ID, ADD_TO_CARD).click()
    self.driver.find_element(By.CLASS_NAME, BASKET_ITEM).click()
    self.driver.find_element(By.ID, PRODUCT_REMOVE).click()
